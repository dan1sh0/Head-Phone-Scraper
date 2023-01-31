#Find price of item and tell me the lowest
# tracks price per day and adds it into a excel file 
#notifies me of lowest price  

#all imports 
from bs4 import BeautifulSoup
import requests 
import numpy as np
import pandas as pd 
import smtplib 
import schedule 
import time 
from openpyxl import load_workbook
from plyer import notification 



link = "https://www.ebay.com/sch/i.html?_from=R40&_trksid=p2334524.m570.l1311&_nkw=sony+wh-1000xm4+headphones+black&_sacat=0&LH_TitleDesc=0&_odkw=sony+xm4+headphones+black&_osacat=0&LH_PrefLoc=2"

def price_finder(link, pages=4):
    item_prices = []
    # find the actual price by parsing through the html, finds the span with the price tag
    # goes through every page
    for page in range(1, pages+1):
        # the pgn is the parameter for pagination for ebay 
        request = requests.get(link + "&_ipg=240&_pgn=" + str(page))
        page_content = BeautifulSoup(request.text, 'html.parser')
        prices = page_content.find("ul",{"class":"srp-results"}).find_all("li",{"class":"s-item"})

        for result in prices:
            price_as_text = result.find("span",{"class":"s-item__price"}).text
            # there are "price to price" on some heaphones and this allows it to keep that in the array 
            if "to" in price_as_text:
               continue
            price = float(price_as_text[1:].replace(",",""))
            item_prices.append(price_as_text)
    return item_prices

        

def remove_outliers(prices):
    # Convert the list of prices from strings to floats
    prices = [float(price[1:].replace(",","")) for price in prices]

    # Calculate the mean and standard deviation of the prices
    mean = np.mean(prices)
    std_dev = np.std(prices)
    
    # Remove any prices more than 1 standard deviations away from the mean
    filtered_prices = [price for price in prices if abs(price - mean) <= 1 * std_dev]
    return filtered_prices


def store_prices_to_excel(prices, filename):
    df = pd.DataFrame(prices, columns=["Price"])
    df.to_excel(filename)



def read_prices_from_excel(filename):
    wb = load_workbook(filename)
    sheet = wb.active
    prices = []
    for row in sheet.iter_rows(values_only=True):
        prices.append(row[0])
    return prices


def notify_lowest_price(prices):
    lowest_price = min(prices)
    notification.notify(
        title='Lowest Price for Headphones',
        message='The lowest price for the headphones is $' + str(lowest_price),
        timeout=10
    )



def notify_me():
    prices = price_finder(link, pages=4)
    filtered_prices = remove_outliers(prices)
    store_prices_to_excel(filtered_prices, 'headphone_prices.xlsx')
    read_prices_from_excel('headphone_prices.xlsx')
    notify_lowest_price(filtered_prices)


if __name__ == "__main__":
    # Schedules the alert to notify me at any point I would like 
    schedule.every().day.at("20:54").do(notify_me)
    while True:
        schedule.run_pending()
        time.sleep(1)


   
    






